"""
Hacker News Scraper & Analyzer

Description:
Scrapes Hacker News posts and extracts:
- title
- link
- author
- time
- score
- comments

Then:
- sorts posts by comments
- prints TOP 5
- saves to a formatted Excel file

Technologies:
- requests
- BeautifulSoup
- openpyxl
"""

from bs4 import BeautifulSoup
import requests
import time
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# Headers to simulate browser
headers = {
    'User-Agent': 'Mozilla/5.0'
}

data = []
pages = 10

# =========================
# SCRAPING
# =========================
for page in range(1, pages + 1):
    time.sleep(1)

    url = f'https://news.ycombinator.com/news?p={page}'
    response = requests.get(url, headers=headers)

    soup = BeautifulSoup(response.text, 'html.parser')
    posts = soup.find_all('tr', class_='athing')

    print(f'Parsing page {page}, posts: {len(posts)}')

    for post in posts:
        # Title + link
        title_tag = post.find("span", class_='titleline').find('a')
        title = title_tag.text
        link = title_tag.get('href')

        if link.startswith('item?id='):
            link = 'https://news.ycombinator.com/' + link

        # Meta
        subtext = post.find_next_sibling("tr")

        score_tag = subtext.find('span', class_='score')
        author_tag = subtext.find('a', class_='hnuser')
        time_tag = subtext.find('span', class_='age')

        score_text = score_tag.text if score_tag else '0 points'
        author = author_tag.text if author_tag else 'unknown'
        post_time = time_tag.text if time_tag else 'no data'

        # Comments
        comment_text = '0 comments'
        for a in subtext.find_all('a'):
            text = a.text.lower()
            if 'comment' in text:
                comment_text = a.text
            elif text == 'discuss':
                comment_text = '0 comments'

        # Convert to numbers
        score = int(score_text.split()[0]) if score_text != '0 points' else 0

        comment_count = 0
        if 'comment' in comment_text:
            comment_count = int(comment_text.split()[0])

        # Save
        data.append({
            'title': title,
            'link': link,
            'author': author,
            'time': post_time,
            'score': score,
            'comments': comment_count
        })

# =========================
# ANALYSIS
# =========================
data = sorted(data, key=lambda x: x['comments'], reverse=True)

print("\n🔥 TOP 5 POSTS BY COMMENTS:\n")
for item in data[:5]:
    print(item)
    print('-' * 50)

# =========================
# EXCEL EXPORT
# =========================
wb = Workbook()
ws = wb.active
ws.title = "Hacker News"

headers_excel = ['title', 'link', 'author', 'time', 'score', 'comments']
ws.append(headers_excel)

# Style header (bold)
for col_num, col_name in enumerate(headers_excel, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = Font(bold=True)

# Add data
for item in data:
    ws.append([
        item['title'],
        item['link'],
        item['author'],
        item['time'],
        item['score'],
        item['comments']
    ])

# Wrap text for all cells
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

# Auto column width
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter

    for cell in column:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass

    ws.column_dimensions[column_letter].width = max_length + 2

# Save file
wb.save('hacker_news.xlsx')

print("\nExcel file saved: hacker_news.xlsx")